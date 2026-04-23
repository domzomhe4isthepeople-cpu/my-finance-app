import streamlit as st
import pandas as pd
import os
from datetime import datetime, date
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ─── CONFIG ────────────────────────────────────────────────────────────────────
FILE_NAME = "money_tracker.xlsx"
NOW = datetime.now()
CURRENT_YEAR  = NOW.year
CURRENT_MONTH = NOW.month
CURRENT_DAY   = NOW.day

MONTH_TH = ["", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
             "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]

CATEGORY_CONFIG = {
    # ประเภท: (emoji, สี, ประเภทหลัก)
    "เงินเดือน":        ("💼", "#22c55e", "รายรับ"),
    "รายได้เสริม":      ("💡", "#10b981", "รายรับ"),
    "โบนัส":            ("🎁", "#34d399", "รายรับ"),
    "ค่าอาหาร":         ("🍜", "#f59e0b", "รายจ่าย"),
    "ค่าเดินทาง":       ("🚗", "#ef4444", "รายจ่าย"),
    "ค่าที่พัก":        ("🏠", "#ec4899", "รายจ่าย"),
    "ค่าสาธารณูปโภค":  ("⚡", "#a78bfa", "รายจ่าย"),
    "ค่าสุขภาพ":        ("🏥", "#06b6d4", "รายจ่าย"),
    "ความบันเทิง":      ("🎮", "#8b5cf6", "รายจ่าย"),
    "ช้อปปิ้ง":         ("🛍️", "#f97316", "รายจ่าย"),
    "การออม":           ("🏦", "#14b8a6", "รายจ่าย"),
    "อื่น ๆ":           ("📌", "#94a3b8", "รายจ่าย"),
}

INCOME_CATS  = [k for k, v in CATEGORY_CONFIG.items() if v[2] == "รายรับ"]
EXPENSE_CATS = [k for k, v in CATEGORY_CONFIG.items() if v[2] == "รายจ่าย"]
ALL_CATS     = INCOME_CATS + EXPENSE_CATS

st.set_page_config(
    page_title="Smart Wealth 2026",
    layout="wide",
    page_icon="💰",
    initial_sidebar_state="expanded",
)

# ─── CUSTOM CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Prompt:wght@300;400;500;600;700&family=Space+Mono&display=swap');

html, body, [class*="css"] {
    font-family: 'Prompt', sans-serif;
}
.stApp { background: #0f172a; color: #e2e8f0; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1e293b 0%, #0f172a 100%);
    border-right: 1px solid #334155;
}

/* Metric cards */
[data-testid="stMetric"] {
    background: #1e293b;
    border: 1px solid #334155;
    border-radius: 16px;
    padding: 20px !important;
}
[data-testid="stMetricLabel"] { color: #94a3b8 !important; font-size: 0.8rem !important; }
[data-testid="stMetricValue"] { color: #f8fafc !important; font-size: 1.6rem !important; font-weight: 700; }
[data-testid="stMetricDelta"] svg { display: none; }

/* Buttons */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
    border: none !important; border-radius: 12px !important;
    font-family: 'Prompt', sans-serif !important; font-weight: 600 !important;
    transition: all .2s !important;
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(99,102,241,.4) !important;
}
.stButton > button[kind="secondary"] {
    background: #1e293b !important; color: #e2e8f0 !important;
    border: 1px solid #475569 !important; border-radius: 12px !important;
    font-family: 'Prompt', sans-serif !important;
}

/* Input / Select */
.stSelectbox > div > div, .stNumberInput > div > div > input,
.stTextInput > div > div > input {
    background: #1e293b !important; border: 1px solid #334155 !important;
    border-radius: 10px !important; color: #f1f5f9 !important;
}

/* Divider */
hr { border-color: #1e293b !important; }

/* Tabs */
button[data-baseweb="tab"] {
    font-family: 'Prompt', sans-serif !important; font-size: .85rem !important;
    color: #94a3b8 !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: #6366f1 !important; border-bottom-color: #6366f1 !important;
}

/* DataEditor */
[data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; }

/* section headers */
.section-header {
    display: flex; align-items: center; gap: 10px;
    font-size: 1.15rem; font-weight: 600; color: #f1f5f9;
    margin-bottom: 1rem;
}

/* badge */
.badge {
    display: inline-block; padding: 2px 10px; border-radius: 20px;
    font-size: .7rem; font-weight: 600; font-family: 'Space Mono', monospace;
}
.badge-income  { background: #166534; color: #4ade80; }
.badge-expense { background: #7f1d1d; color: #f87171; }

/* progress bar */
.budget-bar-wrap { background:#1e293b; border-radius:8px; height:8px; margin:6px 0; }
.budget-bar-fill { height:8px; border-radius:8px; transition: width .4s; }
</style>
""", unsafe_allow_html=True)

# ─── DATA HELPERS ──────────────────────────────────────────────────────────────
def load_data(year: int) -> pd.DataFrame:
    if not os.path.exists(FILE_NAME):
        return pd.DataFrame()
    try:
        df = pd.read_excel(FILE_NAME, sheet_name=str(year))
        df["วันที่"] = pd.to_datetime(df["วันที่"], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()

SETTINGS_SHEET = "_settings"

def _write_sheet(sheet_name: str, df: pd.DataFrame):
    """Safely write a single sheet without touching any other sheets."""
    from openpyxl import load_workbook, Workbook
    if os.path.exists(FILE_NAME):
        wb = load_workbook(FILE_NAME)
    else:
        wb = Workbook()
        # Remove the default empty sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Remove old version of this sheet if it exists
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    # Write the new sheet via a temp ExcelWriter trick using openpyxl directly
    ws = wb.create_sheet(sheet_name)
    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    # Write rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(FILE_NAME)

def save_all_data(df: pd.DataFrame, year: int):
    _write_sheet(str(year), df)

def append_row(row: dict, year: int):
    df = load_data(year)
    new_row = pd.DataFrame([row])
    df = pd.concat([df, new_row], ignore_index=True)
    save_all_data(df, year)

def delete_rows(ids_to_delete: list, year: int):
    df = load_data(year)
    df["ID"] = df["ID"].astype(str)
    ids_str = [str(i) for i in ids_to_delete]
    df = df[~df["ID"].isin(ids_str)]
    save_all_data(df, year)

def load_settings() -> dict:
    if not os.path.exists(FILE_NAME):
        return {}
    try:
        df_s = pd.read_excel(FILE_NAME, sheet_name=SETTINGS_SHEET)
        return dict(zip(df_s["key"], df_s["value"]))
    except Exception:
        return {}

def save_settings(settings: dict):
    df_s = pd.DataFrame(list(settings.items()), columns=["key", "value"])
    _write_sheet(SETTINGS_SHEET, df_s)

# ─── SIDEBAR ───────────────────────────────────────────────────────────────────
# Load persisted settings once per session
_saved = load_settings()

with st.sidebar:
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("## 💰 Smart Wealth")
    st.markdown("<p style='color:#64748b;font-size:.8rem;margin-top:-10px;'>ระบบจัดการการเงินส่วนตัว</p>", unsafe_allow_html=True)
    st.divider()

    selected_year = st.number_input(
        "📅 ปี (ค.ศ.)", value=int(_saved.get("selected_year", CURRENT_YEAR)), step=1, format="%d")
    starting_balance = st.number_input(
        "💵 เงินต้นตั้งต้น (บาท)", min_value=0.0,
        value=float(_saved.get("starting_balance", 0.0)), step=1000.0)

    st.divider()
    st.markdown("**🎯 ตั้งงบประมาณรายเดือน**")
    monthly_budget = st.number_input(
        "งบจ่ายต่อเดือน (บาท)", min_value=0.0,
        value=float(_saved.get("monthly_budget", 0.0)), step=500.0)
    savings_goal = st.number_input(
        "เป้าหมายออม/เดือน (บาท)", min_value=0.0,
        value=float(_saved.get("savings_goal", 0.0)), step=500.0)

    st.divider()
    st.caption("Developed by Chin 🚀  |  v2.1")

# Auto-save settings only when values have actually changed
_current_settings = {
    "selected_year": selected_year,
    "starting_balance": starting_balance,
    "monthly_budget": monthly_budget,
    "savings_goal": savings_goal,
}
_saved_comparable = {k: float(_saved[k]) if k != "selected_year" else int(_saved[k])
                     for k in _current_settings if k in _saved}
_needs_save = any(
    _current_settings[k] != _saved_comparable.get(k)
    for k in _current_settings
)
if _needs_save:
    save_settings(_current_settings)

# ─── LOAD DATA ────────────────────────────────────────────────────────────────
df = load_data(selected_year)

def get_totals(frame):
    if frame.empty:
        return 0.0, 0.0
    inc = frame[frame["ประเภทหลัก"] == "รายรับ"]["จำนวนเงิน"].sum()
    exp = frame[frame["ประเภทหลัก"] == "รายจ่าย"]["จำนวนเงิน"].sum()
    return inc, exp

income_total, expense_total = get_totals(df)
final_balance = starting_balance + income_total - expense_total

# current-month filter
df_month = pd.DataFrame()
if not df.empty:
    df_month = df[df["วันที่"].dt.month == CURRENT_MONTH]
inc_month, exp_month = get_totals(df_month)

# ─── HEADER ───────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style='text-align:center; padding: 1.5rem 0 .5rem;'>
  <h1 style='font-size:2rem; font-weight:700; color:#f1f5f9; letter-spacing:-.5px;'>
    📊 บัญชีการเงิน <span style='color:#6366f1;'>ปี {selected_year}</span>
  </h1>
  <p style='color:#64748b; margin-top:-8px;'>อัปเดตล่าสุด: {NOW.strftime("%d/%m/%Y %H:%M")}</p>
</div>
""", unsafe_allow_html=True)

# ─── KPI CARDS ────────────────────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
balance_delta = income_total - expense_total
c1.metric("💰 เงินคงเหลือ (รวม)", f"฿{final_balance:,.0f}",
          delta=f"{'▲' if balance_delta >= 0 else '▼'} ฿{abs(balance_delta):,.0f}")
c2.metric("📈 รายรับสะสม (ปีนี้)", f"฿{income_total:,.0f}")
c3.metric("📉 รายจ่ายสะสม (ปีนี้)", f"฿{expense_total:,.0f}")
c4.metric(f"🗓️ รายจ่าย {MONTH_TH[CURRENT_MONTH]}", f"฿{exp_month:,.0f}",
          delta=f"รายรับ ฿{inc_month:,.0f}" if inc_month else None)

# Monthly budget progress
if monthly_budget > 0 and exp_month > 0:
    pct = min(exp_month / monthly_budget * 100, 100)
    color = "#22c55e" if pct < 70 else "#f59e0b" if pct < 90 else "#ef4444"
    st.markdown(f"""
    <div style='background:#1e293b;border:1px solid #334155;border-radius:12px;padding:12px 18px;margin:.5rem 0;'>
      <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;'>
        <span style='color:#94a3b8;font-size:.85rem;'>งบประมาณเดือนนี้</span>
        <span style='color:{color};font-weight:700;font-size:.9rem;'>฿{exp_month:,.0f} / ฿{monthly_budget:,.0f} ({pct:.1f}%)</span>
      </div>
      <div class='budget-bar-wrap'><div class='budget-bar-fill' style='width:{pct}%;background:{color};'></div></div>
    </div>""", unsafe_allow_html=True)

st.divider()

# ─── TABS ─────────────────────────────────────────────────────────────────────
tab_overview, tab_add, tab_history, tab_analysis = st.tabs(
    ["📊 ภาพรวม", "➕ บันทึกรายการ", "📋 ประวัติ", "🔍 วิเคราะห์เชิงลึก"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 – OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
with tab_overview:
    if df.empty:
        st.info("ยังไม่มีข้อมูลสำหรับปีนี้ กรุณาเพิ่มรายการในแท็บ 'บันทึกรายการ'")
    else:
        g1, g2 = st.columns([3, 2])

        with g1:
            # Monthly bar chart
            df_graph = df.copy()
            df_graph["เดือน"] = df_graph["วันที่"].dt.month
            monthly = df_graph.groupby(["เดือน", "ประเภทหลัก"])["จำนวนเงิน"].sum().reset_index()
            monthly["ชื่อเดือน"] = monthly["เดือน"].map(lambda m: MONTH_TH[m])

            fig_bar = px.bar(
                monthly, x="ชื่อเดือน", y="จำนวนเงิน", color="ประเภทหลัก",
                barmode="group", title="รายรับ-รายจ่าย รายเดือน",
                color_discrete_map={"รายรับ": "#22c55e", "รายจ่าย": "#f87171"},
                template="plotly_dark",
            )
            fig_bar.update_layout(
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                font_family="Prompt", legend_title_text="",
                title_font_size=15, margin=dict(t=50, b=10, l=0, r=0),
            )
            st.plotly_chart(fig_bar, width='stretch')

        with g2:
            # Expense pie
            exp_df = df[df["ประเภทหลัก"] == "รายจ่าย"]
            if not exp_df.empty:
                fig_pie = px.pie(
                    exp_df.groupby("รายการ")["จำนวนเงิน"].sum().reset_index(),
                    values="จำนวนเงิน", names="รายการ",
                    title="สัดส่วนรายจ่ายตามหมวดหมู่",
                    hole=0.55, template="plotly_dark",
                    color_discrete_sequence=px.colors.qualitative.Pastel,
                )
                fig_pie.update_layout(
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font_family="Prompt", showlegend=True,
                    legend=dict(font_size=11),
                    title_font_size=15, margin=dict(t=50, b=10, l=0, r=0),
                )
                st.plotly_chart(fig_pie, width='stretch')

        # Cumulative balance line
        df_sorted = df.sort_values("วันที่").copy()
        df_sorted["รายรับสุทธิ"] = df_sorted.apply(
            lambda r: r["จำนวนเงิน"] if r["ประเภทหลัก"] == "รายรับ" else -r["จำนวนเงิน"], axis=1)
        df_sorted["ยอดสะสม"] = starting_balance + df_sorted["รายรับสุทธิ"].cumsum()

        fig_line = px.area(
            df_sorted, x="วันที่", y="ยอดสะสม",
            title="ยอดเงินสะสมตลอดปี",
            template="plotly_dark",
            color_discrete_sequence=["#6366f1"],
        )
        fig_line.update_traces(fill="tozeroy", fillcolor="rgba(99,102,241,0.15)")
        fig_line.update_layout(
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font_family="Prompt", title_font_size=15,
            margin=dict(t=50, b=10, l=0, r=0),
        )
        st.plotly_chart(fig_line, width='stretch')

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 – ADD ENTRY
# ══════════════════════════════════════════════════════════════════════════════
with tab_add:
    st.markdown("<div class='section-header'>➕ บันทึกรายการใหม่</div>", unsafe_allow_html=True)

    with st.container(border=True):
        col_type, col_cat = st.columns(2)
        with col_type:
            entry_type = st.radio("ประเภท", ["รายจ่าย", "รายรับ"], horizontal=True)
        with col_cat:
            cats = EXPENSE_CATS if entry_type == "รายจ่าย" else INCOME_CATS
            emoji_cats = [f"{CATEGORY_CONFIG[c][0]} {c}" for c in cats]
            sel_cat_display = st.selectbox("หมวดหมู่", emoji_cats)
            sel_cat = sel_cat_display.split(" ", 1)[1]

        d1, d2, d3 = st.columns([1, 1, 2])
        with d1:
            month = st.selectbox("เดือน", range(1, 13), index=CURRENT_MONTH - 1,
                                  format_func=lambda m: MONTH_TH[m])
        with d2:
            day = st.selectbox("วันที่", range(1, 32), index=CURRENT_DAY - 1)
        with d3:
            amount = st.number_input("จำนวนเงิน (บาท)", min_value=0.0, step=10.0, format="%.2f")

        suggestions = []
        if not df.empty and "หมายเหตุ" in df.columns:
            suggestions = df["หมายเหตุ"].dropna().unique().tolist()

        is_new = st.toggle("พิมพ์รายละเอียดใหม่", value=True)
        if is_new:
            note = st.text_input("รายละเอียด / หมายเหตุ", placeholder="เช่น กะเพราไข่ดาว, ค่าน้ำมัน")
        else:
            note = st.selectbox("เลือกจากประวัติที่เคยบันทึก", options=[""] + suggestions)

        submitted = st.button("💾 บันทึกข้อมูล", width='stretch', type="primary")
        if submitted:
            if amount <= 0:
                st.warning("กรุณาระบุจำนวนเงินมากกว่า 0")
            else:
                row = {
                    "วันที่": f"{selected_year}-{month:02d}-{day:02d}",
                    "รายการ": sel_cat,
                    "ประเภทหลัก": entry_type,
                    "จำนวนเงิน": amount,
                    "หมายเหตุ": note,
                    "ID": NOW.strftime("%Y%m%d%H%M%S%f"),
                }
                append_row(row, selected_year)
                st.success(f"✅ บันทึกสำเร็จ! {'💸' if entry_type=='รายจ่าย' else '💰'} {sel_cat} ฿{amount:,.2f}")
                st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 – HISTORY
# ══════════════════════════════════════════════════════════════════════════════
with tab_history:
    st.markdown("<div class='section-header'>📋 ประวัติรายการทั้งหมด</div>", unsafe_allow_html=True)

    if df.empty:
        st.info("ยังไม่มีข้อมูล")
    else:
        # Filters
        f1, f2, f3 = st.columns(3)
        with f1:
            filter_type = st.selectbox("กรองประเภท", ["ทั้งหมด", "รายรับ", "รายจ่าย"])
        with f2:
            filter_month = st.selectbox("กรองเดือน", ["ทั้งหมด"] + [MONTH_TH[m] for m in range(1, 13)])
        with f3:
            filter_cat = st.selectbox("กรองหมวดหมู่", ["ทั้งหมด"] + sorted(df["รายการ"].dropna().unique().tolist()))

        df_filtered = df.copy()
        if filter_type != "ทั้งหมด":
            df_filtered = df_filtered[df_filtered["ประเภทหลัก"] == filter_type]
        if filter_month != "ทั้งหมด":
            m_idx = MONTH_TH.index(filter_month)
            df_filtered = df_filtered[df_filtered["วันที่"].dt.month == m_idx]
        if filter_cat != "ทั้งหมด":
            df_filtered = df_filtered[df_filtered["รายการ"] == filter_cat]

        st.caption(f"พบ {len(df_filtered):,} รายการ | รายรับ ฿{df_filtered[df_filtered['ประเภทหลัก']=='รายรับ']['จำนวนเงิน'].sum():,.0f} | รายจ่าย ฿{df_filtered[df_filtered['ประเภทหลัก']=='รายจ่าย']['จำนวนเงิน'].sum():,.0f}")

        df_show = df_filtered.copy().sort_values("วันที่", ascending=False)
        if "ลบ" not in df_show.columns:
            df_show.insert(0, "ลบ", False)

        df_show["วันที่_แสดง"] = df_show["วันที่"].dt.strftime("%d/%m/%Y")
        # Force string types to prevent PyArrow OverflowError on large ID integers
        df_show["ID"] = df_show["ID"].astype(str)
        df_show["วันที่_แสดง"] = df_show["วันที่_แสดง"].astype(str)
        df_show["ประเภทหลัก"] = df_show["ประเภทหลัก"].astype(str)
        df_show["รายการ"] = df_show["รายการ"].astype(str)
        df_show["หมายเหตุ"] = df_show["หมายเหตุ"].fillna("").astype(str)
        df_show["จำนวนเงิน"] = pd.to_numeric(df_show["จำนวนเงิน"], errors="coerce").fillna(0.0)
        df_show["ลบ"] = df_show["ลบ"].astype(bool)
        cols_show = ["ลบ", "วันที่_แสดง", "ประเภทหลัก", "รายการ", "จำนวนเงิน", "หมายเหตุ", "ID"]
        df_edit = df_show[[c for c in cols_show if c in df_show.columns]]

        edited = st.data_editor(
            df_edit,
            column_config={
                "ลบ": st.column_config.CheckboxColumn("🗑️", help="เลือกเพื่อลบ"),
                "วันที่_แสดง": st.column_config.TextColumn("วันที่"),
                "ประเภทหลัก": st.column_config.TextColumn("ประเภท"),
                "รายการ": st.column_config.TextColumn("หมวดหมู่"),
                "จำนวนเงิน": st.column_config.NumberColumn("จำนวนเงิน", format="฿%.2f"),
                "หมายเหตุ": st.column_config.TextColumn("หมายเหตุ"),
                # Keep ID visible but read-only so it survives into `edited`
                "ID": st.column_config.TextColumn("ID", disabled=True),
            },
            disabled=["วันที่_แสดง", "ประเภทหลัก", "รายการ", "จำนวนเงิน", "หมายเหตุ", "ID"],
            column_order=["ลบ", "วันที่_แสดง", "ประเภทหลัก", "รายการ", "จำนวนเงิน", "หมายเหตุ"],
            width='stretch',
            hide_index=True,
        )

        btn1, btn2, btn3 = st.columns(3)
        with btn1:
            if st.button("🗑️ ลบรายการที่เลือก", type="secondary", width='stretch'):
                ids_del = edited[edited["ลบ"] == True]["ID"].astype(str).tolist()
                if ids_del:
                    delete_rows(ids_del, selected_year)
                    st.success(f"ลบ {len(ids_del)} รายการเรียบร้อย!")
                    st.rerun()
                else:
                    st.warning("กรุณาติ๊กช่อง 🗑️ ก่อนลบ")
        with btn2:
            csv = df_filtered.to_csv(index=False).encode("utf-8-sig")
            st.download_button("📥 Export CSV", csv, "history.csv", "text/csv", width='stretch')
        with btn3:
            try:
                excel_buf = df_filtered.to_excel.__module__  # trigger import check
                import io
                buf = io.BytesIO()
                df_filtered.to_excel(buf, index=False)
                st.download_button("📊 Export Excel", buf.getvalue(),
                                   "history.xlsx",
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   width='stretch')
            except Exception:
                pass

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 – DEEP ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
with tab_analysis:
    st.markdown("<div class='section-header'>🔍 วิเคราะห์เชิงลึก</div>", unsafe_allow_html=True)

    if df.empty:
        st.info("ยังไม่มีข้อมูลเพียงพอสำหรับการวิเคราะห์")
    else:
        # ── Top spending categories ──
        a1, a2 = st.columns(2)

        with a1:
            st.markdown("**💸 หมวดหมู่ที่ใช้เงินมากที่สุด**")
            exp_by_cat = (
                df[df["ประเภทหลัก"] == "รายจ่าย"]
                .groupby("รายการ")["จำนวนเงิน"].sum()
                .sort_values(ascending=False)
                .reset_index()
            )
            if not exp_by_cat.empty:
                fig_h = px.bar(
                    exp_by_cat.head(8), x="จำนวนเงิน", y="รายการ",
                    orientation="h", template="plotly_dark",
                    color="จำนวนเงิน", color_continuous_scale="Reds",
                )
                fig_h.update_layout(
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font_family="Prompt", showlegend=False,
                    coloraxis_showscale=False,
                    margin=dict(t=10, b=10, l=0, r=0), yaxis_title="", xaxis_title="บาท",
                )
                st.plotly_chart(fig_h, width='stretch')

        with a2:
            st.markdown("**📅 รายจ่ายรายวัน (30 วันล่าสุด)**")
            df_daily = df[df["ประเภทหลัก"] == "รายจ่าย"].copy()
            if not df_daily.empty:
                df_daily["วันที่"] = pd.to_datetime(df_daily["วันที่"])
                daily = df_daily.groupby("วันที่")["จำนวนเงิน"].sum().reset_index().tail(30)
                fig_daily = px.line(
                    daily, x="วันที่", y="จำนวนเงิน",
                    template="plotly_dark", markers=True,
                    color_discrete_sequence=["#f87171"],
                )
                fig_daily.update_traces(marker_size=5)
                fig_daily.update_layout(
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font_family="Prompt",
                    margin=dict(t=10, b=10, l=0, r=0),
                )
                st.plotly_chart(fig_daily, width='stretch')

        # ── Monthly savings rate ──
        st.markdown("**💹 อัตราการออมรายเดือน**")
        df_m = df.copy()
        df_m["เดือน"] = df_m["วันที่"].dt.month
        m_inc = df_m[df_m["ประเภทหลัก"] == "รายรับ"].groupby("เดือน")["จำนวนเงิน"].sum()
        m_exp = df_m[df_m["ประเภทหลัก"] == "รายจ่าย"].groupby("เดือน")["จำนวนเงิน"].sum()
        m_df  = pd.DataFrame({"รายรับ": m_inc, "รายจ่าย": m_exp}).fillna(0)
        m_df["ออมได้"] = m_df["รายรับ"] - m_df["รายจ่าย"]
        m_df["อัตราออม%"] = (m_df["ออมได้"] / m_df["รายรับ"].replace(0, 1) * 100).round(1)
        m_df.index = m_df.index.map(lambda m: MONTH_TH[m])
        m_df = m_df.reset_index().rename(columns={"index": "เดือน"})

        fig_save = go.Figure()
        fig_save.add_trace(go.Bar(
            x=m_df["เดือน"], y=m_df["ออมได้"],
            marker_color=["#22c55e" if v >= 0 else "#ef4444" for v in m_df["ออมได้"]],
            name="ออมได้ (บาท)",
        ))
        fig_save.add_trace(go.Scatter(
            x=m_df["เดือน"], y=m_df["อัตราออม%"],
            mode="lines+markers+text", name="อัตราออม (%)",
            yaxis="y2", line_color="#6366f1",
            text=m_df["อัตราออม%"].map(lambda v: f"{v}%"),
            textposition="top center", textfont_size=10,
        ))
        fig_save.update_layout(
            yaxis2=dict(overlaying="y", side="right", title="อัตราออม (%)"),
            template="plotly_dark",
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font_family="Prompt", legend=dict(orientation="h"),
            margin=dict(t=20, b=10, l=0, r=0),
        )
        st.plotly_chart(fig_save, width='stretch')

        # ── Summary table ──
        st.markdown("**📊 สรุปรายเดือน**")
        st.dataframe(
            m_df.style.format({
                "รายรับ": "฿{:,.0f}", "รายจ่าย": "฿{:,.0f}",
                "ออมได้": "฿{:,.0f}", "อัตราออม%": "{:.1f}%",
            }).map(
                lambda v: "color:#4ade80" if isinstance(v, (int, float)) and v >= 0 else "color:#f87171",
                subset=["ออมได้"],
            ),
            width='stretch', hide_index=True,
        )
